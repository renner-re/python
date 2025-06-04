from openpyxl import load_workbook
from pathlib import Path

def split_names(input_file, name_column='A', start_row=2):
    """
    split full name into first/last
    parameters: 
        input_file = path to xlsx
        name_column = column letter containing full names
        start_row = first row containing data (default 2)
    """
    # load workbook and select active sheet
    wb = load_workbook(input_file)
    sheet = wb.active
    
    # convert column letter to column index
    from openpyxl.utils import column_index_from_string
    name_col_idx = column_index_from_string(name_column)

    # add headers for new columns
    first_name_col = name_col_idx
    last_name_col = name_col_idx + 1

    sheet.cell(row=1, column=first_name_col, value='First Name')
    sheet.cell(row=1, column=last_name_col, value='Last Name')

    # process rows
    for row in range(start_row, sheet.max_row + 1):
        full_name = sheet.cell(row=row, column=name_col_idx).value
        if full_name:
            name_parts = full_name.strip().split(maxsplit=1)

            first_name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ''

            sheet.cell(row=row, column=first_name_col, value=first_name)
            sheet.cell(row=row, column=last_name_col, value=last_name)
    
    wb.save(input_file)
    print(f"Names split succesfully, {input_file} updated.")

if __name__ == "__main__":
    input_file = "/Users/reed/Desktop/users.xlsx"
    split_names(input_file)

